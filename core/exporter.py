from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from core.models import Invoice, InvoiceSheet, InvoiceStatus

SPECIAL_NORMAL_HEADERS = [
    "序号", "发票种类", "发票号码", "开票日期",
    "货物/服务名称", "销方名称",
    "金额", "税率", "税额", "价税合计",
]
MISC_HEADERS = ["序号", "日期", "货物/服务名称", "价税合计"]
EXPORT_STATUSES = {InvoiceStatus.CONFIRMED, InvoiceStatus.MANUAL_DONE}


def _write_header(ws, headers: list):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEEFF")
        cell.alignment = Alignment(horizontal="center")


def export_to_excel(invoices: List[Invoice], output_path: Path):
    wb = Workbook()
    ws_special = wb.active
    ws_special.title = "专票"
    ws_normal = wb.create_sheet("普票")
    ws_misc = wb.create_sheet("杂票")

    for ws in (ws_special, ws_normal):
        _write_header(ws, SPECIAL_NORMAL_HEADERS)
    _write_header(ws_misc, MISC_HEADERS)

    def _sort_key(inv: Invoice) -> str:
        """将开票日期统一转为 YYYY-MM-DD 字符串用于排序，解析失败的排到最后。"""
        import re
        d = inv.issue_date or ""
        # 格式1: "2026年04月14日"
        m = re.match(r"(\d{4})年(\d{2})月(\d{2})日", d)
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        # 格式2: "2026-04-14"
        if re.match(r"\d{4}-\d{2}-\d{2}", d):
            return d
        return "9999-99-99"

    sorted_invoices = sorted(
        [inv for inv in invoices if inv.status in EXPORT_STATUSES],
        key=_sort_key,
    )

    counters = {"专票": 1, "普票": 1, "杂票": 1}

    for inv in sorted_invoices:
        sheet_name = inv.sheet.value
        idx = counters[sheet_name]
        counters[sheet_name] += 1

        if inv.sheet == InvoiceSheet.SPECIAL:
            ws = ws_special
        elif inv.sheet == InvoiceSheet.NORMAL:
            ws = ws_normal
        else:
            ws = ws_misc

        row = ws.max_row + 1
        if inv.sheet in (InvoiceSheet.SPECIAL, InvoiceSheet.NORMAL):
            for col, val in enumerate([
                idx, inv.invoice_type, inv.invoice_number,
                inv.issue_date, inv.goods_name, inv.seller_name,
                inv.amount, inv.tax_rate, inv.tax_amount, inv.total_amount,
            ], 1):
                ws.cell(row, col, val)
        else:
            for col, val in enumerate([
                idx, inv.issue_date, inv.goods_name, inv.total_amount
            ], 1):
                ws.cell(row, col, val)

    wb.save(output_path)

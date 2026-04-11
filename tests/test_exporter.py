import tempfile
from pathlib import Path
from openpyxl import load_workbook
from core.exporter import export_to_excel
from core.models import Invoice, InvoiceStatus, InvoiceSheet


def _make_invoice(sheet, number, total) -> Invoice:
    return Invoice(
        file_path=f"/tmp/{number}.pdf",
        status=InvoiceStatus.CONFIRMED,
        sheet=sheet,
        invoice_type="增值税专用发票" if sheet == InvoiceSheet.SPECIAL else "增值税普通发票",
        invoice_code="044031900111", invoice_number=number,
        issue_date="2024年01月01日", goods_name="办公用品",
        seller_name="卖方", buyer_name="买方", buyer_tax_id="91310000XX",
        amount=round(total / 1.13, 2), tax_rate="13%",
        tax_amount=round(total - total / 1.13, 2),
        total_amount=total,
    )


def test_export_creates_three_sheets():
    invoices = [
        _make_invoice(InvoiceSheet.SPECIAL, "00000001", 1130.0),
        _make_invoice(InvoiceSheet.NORMAL, "00000002", 100.0),
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "output.xlsx"
        export_to_excel(invoices, out)
        wb = load_workbook(out)
        assert "专票" in wb.sheetnames
        assert "普票" in wb.sheetnames
        assert "杂票" in wb.sheetnames


def test_export_special_sheet_columns():
    invoices = [_make_invoice(InvoiceSheet.SPECIAL, "00000001", 1130.0)]
    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "output.xlsx"
        export_to_excel(invoices, out)
        wb = load_workbook(out)
        ws = wb["专票"]
        headers = [ws.cell(1, c).value for c in range(1, 14)]
        assert headers[0] == "序号"
        assert "价税合计" in headers


def test_export_misc_sheet_columns():
    inv = _make_invoice(InvoiceSheet.MISC, "00000003", 50.0)
    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "output.xlsx"
        export_to_excel([inv], out)
        wb = load_workbook(out)
        ws = wb["杂票"]
        headers = [ws.cell(1, c).value for c in range(1, 5)]
        assert headers == ["序号", "日期", "货物/服务名称", "价税合计"]


def test_only_confirmed_and_manual_done_exported():
    inv_confirmed = _make_invoice(InvoiceSheet.NORMAL, "00000001", 100.0)
    inv_pending = _make_invoice(InvoiceSheet.NORMAL, "00000002", 200.0)
    inv_pending.status = InvoiceStatus.PENDING
    inv_manual = _make_invoice(InvoiceSheet.NORMAL, "00000003", 300.0)
    inv_manual.status = InvoiceStatus.MANUAL_DONE
    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "output.xlsx"
        export_to_excel([inv_confirmed, inv_pending, inv_manual], out)
        wb = load_workbook(out)
        ws = wb["普票"]
        assert ws.max_row == 3  # header + 2 data rows
